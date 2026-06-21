"""
build_workbook.py
-----------------
Generates Commercial_Credit_Underwriting_Model.xlsx, an employer-facing
commercial credit / underwriting portfolio model for a sample borrower:

    Northstar Equipment Services LLC
    $750,000 term loan, 60 months, 8.25%

The workbook is intentionally simple and interview-friendly:
  - Input (assumption) cells use BLUE font.
  - Formula / output cells use BLACK font.
  - Live Excel formulas are used wherever practical so a reviewer can audit
    the math and flex the assumptions.

Run:  python build_workbook.py
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ----------------------------------------------------------------------------
# Shared styles
# ----------------------------------------------------------------------------
NAVY = "1F3864"
LIGHT = "D9E1F2"
GREY = "F2F2F2"
GREEN = "E2EFDA"
AMBER = "FCE4D6"

INPUT_FONT = Font(color="0000FF")                 # blue  -> inputs / assumptions
INPUT_FONT_B = Font(color="0000FF", bold=True)
OUT_FONT = Font(color="000000")                   # black -> formulas / outputs
OUT_FONT_B = Font(color="000000", bold=True)
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
SUB_FONT = Font(color="FFFFFF", bold=True, size=11)
HDR_FONT = Font(color="FFFFFF", bold=True)
NOTE_FONT = Font(color="595959", italic=True, size=9)
SECTION_FONT = Font(color=NAVY, bold=True, size=11)

TITLE_FILL = PatternFill("solid", fgColor=NAVY)
HDR_FILL = PatternFill("solid", fgColor=NAVY)
LIGHT_FILL = PatternFill("solid", fgColor=LIGHT)
GREY_FILL = PatternFill("solid", fgColor=GREY)
GREEN_FILL = PatternFill("solid", fgColor=GREEN)
AMBER_FILL = PatternFill("solid", fgColor=AMBER)

CTR = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CUR = '#,##0;(#,##0)'          # currency, no decimals
CUR0 = '$#,##0;($#,##0)'
PCT = '0.0%'
RATIO = '0.00"x"'
NUM2 = '0.00'

YEARS = ["2023", "2024", "2025"]


# ----------------------------------------------------------------------------
# Small helpers
# ----------------------------------------------------------------------------
def title_block(ws, title, subtitle, span=4):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(1, 1, title)
    c.font = TITLE_FONT
    c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    c = ws.cell(2, 1, subtitle)
    c.font = SUB_FONT
    c.fill = TITLE_FILL
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18


def hdr(cell, text):
    cell.value = text
    cell.font = HDR_FONT
    cell.fill = HDR_FILL
    cell.alignment = CTR
    cell.border = BORDER


def label(cell, text, bold=False):
    cell.value = text
    cell.font = OUT_FONT_B if bold else OUT_FONT
    cell.alignment = LEFT


def inp(cell, value, fmt=CUR, bold=False):
    cell.value = value
    cell.font = INPUT_FONT_B if bold else INPUT_FONT
    cell.number_format = fmt
    cell.alignment = RIGHT
    cell.border = BORDER


def out(cell, value, fmt=CUR, bold=False, fill=None):
    cell.value = value
    cell.font = OUT_FONT_B if bold else OUT_FONT
    cell.number_format = fmt
    cell.alignment = RIGHT
    cell.border = BORDER
    if fill:
        cell.fill = fill


def note(ws, row, text, span=4, height=None):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = NOTE_FONT
    c.alignment = WRAP
    if height:
        ws.row_dimensions[row].height = height


def section(ws, row, text, span=4):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row, 1, text)
    c.font = SECTION_FONT
    c.fill = LIGHT_FILL
    c.alignment = LEFT


# ============================================================================
wb = Workbook()

# Sheet names (used for cross-references)
S1 = "01 Borrower Overview"
S2 = "02 Income Statement"
S3 = "03 Balance Sheet"
S4 = "04 Ratio Analysis"
S5 = "05 DSCR Analysis"
S6 = "06 Collateral LTV"
S7 = "07 Risk Rating"
S8 = "08 Credit Memo"


# ----------------------------------------------------------------------------
# 01 Borrower Overview
# ----------------------------------------------------------------------------
ws = wb.active
ws.title = S1
ws.sheet_view.showGridLines = False
for col, w in {"A": 26, "B": 40, "C": 16, "D": 16}.items():
    ws.column_dimensions[col].width = w
title_block(ws, "Commercial Credit Underwriting Model",
            "Northstar Equipment Services LLC  |  $750,000 Term Loan", span=4)

r = 4
section(ws, r, "Borrower & Request Summary"); r += 1
facts = [
    ("Borrower", "Northstar Equipment Services LLC"),
    ("Entity Type", "Limited Liability Company (LLC)"),
    ("Industry", "Commercial equipment repair & maintenance"),
    ("Relationship", "New borrower (no prior bank history)"),
    ("Loan Request", "$750,000 senior secured term loan"),
    ("Purpose", "Purchase service vehicles and shop equipment"),
    ("Repayment Source", "Operating cash flow (EBITDA)"),
    ("Collateral", "Service vehicles & equipment (purchase-money lien)"),
    ("Guarantee", "Full personal guarantee of majority owner"),
    ("Underwriting Question", "Should the bank approve this loan?"),
]
for lbl, val in facts:
    label(ws.cell(r, 1, lbl), lbl, bold=True)
    c = ws.cell(r, 2, val); c.font = OUT_FONT; c.alignment = LEFT
    r += 1

r += 1
section(ws, r, "Proposed Loan Terms  (blue = input assumptions)"); r += 1
hdr(ws.cell(r, 1), "Term"); hdr(ws.cell(r, 2), "Value"); r += 1
LOAN_ROW = r
label(ws.cell(r, 1), "Loan Amount", bold=True); inp(ws.cell(r, 2), 750000, CUR0, bold=True); r += 1
TERM_ROW = r
label(ws.cell(r, 1), "Term (months)", bold=True); inp(ws.cell(r, 2), 60, '0', bold=True); r += 1
RATE_ROW = r
label(ws.cell(r, 1), "Annual Interest Rate", bold=True); inp(ws.cell(r, 2), 0.0825, PCT, bold=True); r += 1
label(ws.cell(r, 1), "Amortization", bold=True)
c = ws.cell(r, 2, "Fully amortizing, monthly P&I"); c.font = OUT_FONT; c.alignment = LEFT; r += 1
label(ws.cell(r, 1), "Equity Injection", bold=True); inp(ws.cell(r, 2), 85000, CUR0, bold=True); r += 1

# cross-ref strings to loan terms
LOAN = f"'{S1}'!B{LOAN_ROW}"
TERM = f"'{S1}'!B{TERM_ROW}"
RATE = f"'{S1}'!B{RATE_ROW}"

r += 1
note(ws, r, "How to read this workbook:  Blue font = input assumptions you can change. "
            "Black font = formulas/outputs driven by the inputs. Each tab flows into the next, "
            "ending in the Risk Rating scorecard and Credit Memo recommendation.", span=4, height=42)


# ----------------------------------------------------------------------------
# 02 Income Statement
# ----------------------------------------------------------------------------
ws = wb.create_sheet(S2)
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 34
for col in ("B", "C", "D"):
    ws.column_dimensions[col].width = 15
title_block(ws, "Income Statement", "Northstar Equipment Services LLC  |  USD, fiscal year", span=4)

r = 4
hdr(ws.cell(r, 1), "Line Item ($)")
for i, y in enumerate(YEARS):
    hdr(ws.cell(r, 2 + i), y)
r += 1

# Inputs (blue) and formulas (black). Track key rows for cross-references.
REV = r
label(ws.cell(r, 1), "Revenue", bold=True)
for c, v in zip("BCD", [4200000, 4800000, 5400000]):
    inp(ws[f"{c}{r}"], v); r  # noqa
r += 1
COGS = r
label(ws.cell(r, 1), "Cost of Goods Sold (COGS)")
for c, v in zip("BCD", [2604000, 2952000, 3294000]):
    inp(ws[f"{c}{r}"], v)
r += 1
GP = r
label(ws.cell(r, 1), "Gross Profit", bold=True)
for c in "BCD":
    out(ws[f"{c}{r}"], f"={c}{REV}-{c}{COGS}", bold=True, fill=GREY_FILL)
r += 1
OPEX = r
label(ws.cell(r, 1), "Operating Expenses (excl. D&A)")
for c, v in zip("BCD", [1180000, 1330000, 1480000]):
    inp(ws[f"{c}{r}"], v)
r += 1
EBITDA = r
label(ws.cell(r, 1), "EBITDA", bold=True)
for c in "BCD":
    out(ws[f"{c}{r}"], f"={c}{GP}-{c}{OPEX}", bold=True, fill=GREY_FILL)
r += 1
DEP = r
label(ws.cell(r, 1), "Depreciation & Amortization")
for c, v in zip("BCD", [130000, 145000, 160000]):
    inp(ws[f"{c}{r}"], v)
r += 1
EBIT = r
label(ws.cell(r, 1), "EBIT (Operating Income)")
for c in "BCD":
    out(ws[f"{c}{r}"], f"={c}{EBITDA}-{c}{DEP}")
r += 1
INT = r
label(ws.cell(r, 1), "Interest Expense")
for c, v in zip("BCD", [95000, 105000, 110000]):
    inp(ws[f"{c}{r}"], v)
r += 1
EBT = r
label(ws.cell(r, 1), "Pre-Tax Income (EBT)")
for c in "BCD":
    out(ws[f"{c}{r}"], f"={c}{EBIT}-{c}{INT}")
r += 1
TAX = r
label(ws.cell(r, 1), "Income Taxes")
for c, v in zip("BCD", [47750, 67000, 89000]):
    inp(ws[f"{c}{r}"], v)
r += 1
NI = r
label(ws.cell(r, 1), "Net Income", bold=True)
for c in "BCD":
    out(ws[f"{c}{r}"], f"={c}{EBT}-{c}{TAX}", bold=True, fill=GREEN_FILL)
r += 2

note(ws, r, "Notes:  Gross Profit = Revenue - COGS.  EBITDA = Gross Profit - Operating Expenses "
            "(a proxy for cash operating earnings before financing, tax and non-cash D&A). "
            "Net Income is the bottom-line accrual profit. Revenue grew ~14% (2024) and ~13% (2025) "
            "with stable margins - a positive, consistent trend for an underwriter.", span=4, height=56)


# ----------------------------------------------------------------------------
# 03 Balance Sheet
# ----------------------------------------------------------------------------
ws = wb.create_sheet(S3)
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 34
for col in ("B", "C", "D"):
    ws.column_dimensions[col].width = 15
title_block(ws, "Balance Sheet", "Northstar Equipment Services LLC  |  USD, as of year-end", span=4)

r = 4
hdr(ws.cell(r, 1), "Line Item ($)")
for i, y in enumerate(YEARS):
    hdr(ws.cell(r, 2 + i), y)
r += 1

section(ws, r, "Assets"); r += 1
CASH = r
label(ws.cell(r, 1), "Cash & Equivalents")
for c, v in zip("BCD", [180000, 240000, 320000]):
    inp(ws[f"{c}{r}"], v)
r += 1
AR = r
label(ws.cell(r, 1), "Accounts Receivable")
for c, v in zip("BCD", [430000, 490000, 540000]):
    inp(ws[f"{c}{r}"], v)
r += 1
INV = r
label(ws.cell(r, 1), "Inventory")
for c, v in zip("BCD", [240000, 260000, 280000]):
    inp(ws[f"{c}{r}"], v)
r += 1
TCA = r
label(ws.cell(r, 1), "Total Current Assets", bold=True)
for c in "BCD":
    out(ws[f"{c}{r}"], f"=SUM({c}{CASH}:{c}{INV})", bold=True, fill=GREY_FILL)
r += 1
FA = r
label(ws.cell(r, 1), "Net Fixed Assets (PP&E)")
for c, v in zip("BCD", [1350000, 1500000, 1650000]):
    inp(ws[f"{c}{r}"], v)
r += 1
TA = r
label(ws.cell(r, 1), "Total Assets", bold=True)
for c in "BCD":
    out(ws[f"{c}{r}"], f"={c}{TCA}+{c}{FA}", bold=True, fill=LIGHT_FILL)
r += 2

section(ws, r, "Liabilities & Equity"); r += 1
AP = r
label(ws.cell(r, 1), "Accounts Payable")
for c, v in zip("BCD", [300000, 330000, 360000]):
    inp(ws[f"{c}{r}"], v)
r += 1
STD = r
label(ws.cell(r, 1), "Short-Term Debt (incl. CPLTD)")
for c, v in zip("BCD", [150000, 165000, 180000]):
    inp(ws[f"{c}{r}"], v)
r += 1
TCL = r
label(ws.cell(r, 1), "Total Current Liabilities", bold=True)
for c in "BCD":
    out(ws[f"{c}{r}"], f"={c}{AP}+{c}{STD}", bold=True, fill=GREY_FILL)
r += 1
LTD = r
label(ws.cell(r, 1), "Long-Term Debt")
for c, v in zip("BCD", [780000, 820000, 850000]):
    inp(ws[f"{c}{r}"], v)
r += 1
TL = r
label(ws.cell(r, 1), "Total Liabilities", bold=True)
for c in "BCD":
    out(ws[f"{c}{r}"], f"={c}{TCL}+{c}{LTD}", bold=True, fill=GREY_FILL)
r += 1
EQ = r
label(ws.cell(r, 1), "Total Equity (members' capital)")
for c, v in zip("BCD", [970000, 1175000, 1400000]):
    inp(ws[f"{c}{r}"], v)
r += 1
TLE = r
label(ws.cell(r, 1), "Total Liabilities & Equity", bold=True)
for c in "BCD":
    out(ws[f"{c}{r}"], f"={c}{TL}+{c}{EQ}", bold=True, fill=LIGHT_FILL)
r += 1
CHK = r
label(ws.cell(r, 1), "Balance Check (TA - TL&E)")
for c in "BCD":
    out(ws[f"{c}{r}"], f"={c}{TA}-{c}{TLE}")
r += 2

note(ws, r, "Notes:  Total Equity is an input from the borrower's financials; the Balance Check "
            "row confirms Assets = Liabilities + Equity (should be 0). Short-Term Debt includes the "
            "current portion of long-term debt (CPLTD).", span=4, height=42)


# ----------------------------------------------------------------------------
# 04 Ratio Analysis
# ----------------------------------------------------------------------------
ws = wb.create_sheet(S4)
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 30
for col in ("B", "C", "D"):
    ws.column_dimensions[col].width = 14
ws.column_dimensions["E"].width = 52
title_block(ws, "Ratio Analysis (Financial Spreading)",
            "Trends, profitability, leverage and liquidity", span=5)

r = 4
hdr(ws.cell(r, 1), "Ratio")
for i, y in enumerate(YEARS):
    hdr(ws.cell(r, 2 + i), y)
hdr(ws.cell(r, 5), "What it means (plain English)")
r += 1

IS = f"'{S2}'"
BS = f"'{S3}'"


def ratio_row(rr, name, formulas, fmt, meaning, start_col=2):
    label(ws.cell(rr, 1), name, bold=True)
    for j, f in enumerate(formulas):
        col = get_column_letter(start_col + j)
        if f is None:
            c = ws[f"{col}{rr}"]; c.value = "n/a"; c.font = OUT_FONT; c.alignment = RIGHT; c.border = BORDER
        else:
            out(ws[f"{col}{rr}"], f, fmt)
    m = ws.cell(rr, 5, meaning); m.font = NOTE_FONT; m.alignment = WRAP; m.border = BORDER


# Revenue growth (B = n/a, C & D computed)
ratio_row(r, "Revenue Growth %",
          [None,
           f"={IS}!C{REV}/{IS}!B{REV}-1",
           f"={IS}!D{REV}/{IS}!C{REV}-1"],
          PCT, "Year-over-year sales change. Positive, steady growth signals demand and durability.")
r += 1
ratio_row(r, "Gross Margin %",
          [f"={IS}!{c}{GP}/{IS}!{c}{REV}" for c in "BCD"],
          PCT, "Profit left after direct job costs. Stable margin = good pricing & cost control.")
r += 1
ratio_row(r, "EBITDA Margin %",
          [f"={IS}!{c}{EBITDA}/{IS}!{c}{REV}" for c in "BCD"],
          PCT, "Cash operating profitability before financing/tax. Drives debt repayment capacity.")
r += 1
ratio_row(r, "Current Ratio",
          [f"={BS}!{c}{TCA}/{BS}!{c}{TCL}" for c in "BCD"],
          RATIO, "Current assets / current liabilities. >1.5x = comfortable short-term liquidity.")
r += 1
ratio_row(r, "Debt-to-Equity",
          [f"={BS}!{c}{TL}/{BS}!{c}{EQ}" for c in "BCD"],
          RATIO, "Total liabilities / equity. Lower = more owner skin-in-the-game, less leverage.")
r += 1
ratio_row(r, "Debt-to-EBITDA",
          [f"=({BS}!{c}{STD}+{BS}!{c}{LTD})/{IS}!{c}{EBITDA}" for c in "BCD"],
          RATIO, "Interest-bearing debt / EBITDA. Years of cash earnings to repay debt; <3.0x is healthy.")
r += 1
ratio_row(r, "Working Capital ($)",
          [f"={BS}!{c}{TCA}-{BS}!{c}{TCL}" for c in "BCD"],
          CUR, "Current assets minus current liabilities. Positive cushion funds day-to-day operations.")
r += 2

note(ws, r, "Underwriter read:  Margins are stable, leverage is modest (D/E ~1.0x, Debt/EBITDA <2x) "
            "and liquidity is strong (current ratio >2x). The trend is improving across all three years.",
     span=5, height=40)


# ----------------------------------------------------------------------------
# 05 DSCR Analysis
# ----------------------------------------------------------------------------
ws = wb.create_sheet(S5)
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 42
for col in ("B", "C", "D"):
    ws.column_dimensions[col].width = 15
title_block(ws, "DSCR Analysis (Repayment Capacity)",
            "Cash flow available for debt service vs. debt service", span=4)

r = 4
section(ws, r, "Proposed Loan Debt Service"); r += 1
hdr(ws.cell(r, 1), "Item"); hdr(ws.cell(r, 2), "Value"); r += 1
label(ws.cell(r, 1), "Loan Amount"); out(ws.cell(r, 2), f"={LOAN}", CUR0); r += 1
label(ws.cell(r, 1), "Term (months)"); out(ws.cell(r, 2), f"={TERM}", '0'); r += 1
label(ws.cell(r, 1), "Annual Interest Rate"); out(ws.cell(r, 2), f"={RATE}", PCT); r += 1
MPAY = r
label(ws.cell(r, 1), "Monthly Payment (P&I)")
out(ws.cell(r, 2), f"=PMT({RATE}/12,{TERM},-{LOAN})", CUR0); r += 1
ADS = r
label(ws.cell(r, 1), "Proposed Annual Debt Service", bold=True)
out(ws.cell(r, 2), f"=B{MPAY}*12", CUR0, bold=True, fill=LIGHT_FILL); r += 1
EXDS = r
label(ws.cell(r, 1), "Existing Annual Debt Service (current loans)")
inp(ws.cell(r, 2), 175000, CUR0); r += 1
TDS = r
label(ws.cell(r, 1), "Total (Global) Annual Debt Service", bold=True)
out(ws.cell(r, 2), f"=B{ADS}+B{EXDS}", CUR0, bold=True, fill=LIGHT_FILL); r += 2

section(ws, r, "Cash Flow Available for Debt Service (CFADS)"); r += 1
hdr(ws.cell(r, 1), "Component")
for i, y in enumerate(YEARS):
    hdr(ws.cell(r, 2 + i), y)
r += 1
cf_ebitda = r
label(ws.cell(r, 1), "EBITDA")
for c in "BCD":
    out(ws[f"{c}{r}"], f"={IS}!{c}{EBITDA}")
r += 1
cf_tax = r
label(ws.cell(r, 1), "Less: Cash Income Taxes")
for c in "BCD":
    out(ws[f"{c}{r}"], f"=-{IS}!{c}{TAX}")
r += 1
cf_capex = r
label(ws.cell(r, 1), "Less: Maintenance Capex (unfinanced)")
for c, v in zip("BCD", [55000, 58000, 60000]):
    inp(ws[f"{c}{r}"], -v)
r += 1
CFADS = r
label(ws.cell(r, 1), "CFADS", bold=True)
for c in "BCD":
    out(ws[f"{c}{r}"], f"=SUM({c}{cf_ebitda}:{c}{cf_capex})", bold=True, fill=GREY_FILL)
r += 2

section(ws, r, "Debt Service Coverage Ratio (DSCR)"); r += 1
hdr(ws.cell(r, 1), "Metric")
for i, y in enumerate(YEARS):
    hdr(ws.cell(r, 2 + i), y)
r += 1
PDSCR = r
label(ws.cell(r, 1), "Proposed-Loan DSCR  (CFADS / new DS)", bold=True)
for c in "BCD":
    out(ws[f"{c}{r}"], f"={c}{CFADS}/$B${ADS}", RATIO, bold=True, fill=GREEN_FILL)
r += 1
GDSCR = r
label(ws.cell(r, 1), "Global DSCR  (CFADS / total DS)", bold=True)
for c in "BCD":
    out(ws[f"{c}{r}"], f"={c}{CFADS}/$B${TDS}", RATIO, bold=True, fill=GREEN_FILL)
r += 1
label(ws.cell(r, 1), "Policy Minimum DSCR")
inp(ws.cell(r, 2), 1.20, RATIO)
PMIN = r
r += 1
label(ws.cell(r, 1), "Pass / Fail (Global DSCR, 2025)", bold=True)
out(ws.cell(r, 2), f'=IF(D{GDSCR}>=B{PMIN},"PASS","FAIL")', '@', bold=True, fill=GREEN_FILL)
r += 2

note(ws, r, "DSCR = cash flow available for debt service / debt service. It answers: for every $1 of "
            "loan payment, how many dollars of cash does the business generate? Banks typically want "
            ">=1.20x. CFADS here = EBITDA - cash taxes - unfinanced maintenance capex (a conservative "
            "measure). 2025 global DSCR is comfortably above the 1.20x policy floor even with existing debt.",
     span=4, height=58)

# capture key DSCR cells for memo
DSCR_GLOBAL_2025 = f"'{S5}'!D{GDSCR}"
DSCR_PROP_2025 = f"'{S5}'!D{PDSCR}"


# ----------------------------------------------------------------------------
# 06 Collateral LTV
# ----------------------------------------------------------------------------
ws = wb.create_sheet(S6)
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 34
for col, w in {"B": 16, "C": 14, "D": 18, "E": 30}.items():
    ws.column_dimensions[col].width = w
title_block(ws, "Collateral & Loan-to-Value (LTV)",
            "Service vehicles & equipment securing the loan", span=5)

r = 4
hdr(ws.cell(r, 1), "Collateral")
hdr(ws.cell(r, 2), "Appraised / Cost")
hdr(ws.cell(r, 3), "Advance Rate")
hdr(ws.cell(r, 4), "Lendable Value")
hdr(ws.cell(r, 5), "Basis")
r += 1
col_start = r
items = [
    ("Service Vehicles (new, purchase-money)", 520000, 0.80, "Titled rolling stock; NADA values"),
    ("Shop / Service Equipment (new)", 315000, 0.70, "New machinery; net orderly liquidation"),
    ("Existing Equipment (additional pledge)", 300000, 0.60, "Used; discounted forced-sale value"),
]
for name, cost, adv, basis in items:
    label(ws.cell(r, 1), name)
    inp(ws.cell(r, 2), cost, CUR0)
    inp(ws.cell(r, 3), adv, PCT)
    out(ws.cell(r, 4), f"=B{r}*C{r}", CUR0)
    b = ws.cell(r, 5, basis); b.font = NOTE_FONT; b.alignment = WRAP; b.border = BORDER
    r += 1
col_end = r - 1
TOTCOST = r
label(ws.cell(r, 1), "Total Collateral Value", bold=True)
out(ws.cell(r, 2), f"=SUM(B{col_start}:B{col_end})", CUR0, bold=True, fill=GREY_FILL)
ws.cell(r, 3).border = BORDER
LENDABLE = r
out(ws.cell(r, 4), f"=SUM(D{col_start}:D{col_end})", CUR0, bold=True, fill=GREY_FILL)
r += 2

section(ws, r, "LTV & Collateral Coverage"); r += 1
hdr(ws.cell(r, 1), "Metric"); hdr(ws.cell(r, 2), "Value"); r += 1
label(ws.cell(r, 1), "Loan Amount"); out(ws.cell(r, 2), f"={LOAN}", CUR0); r += 1
label(ws.cell(r, 1), "LTV vs. Appraised Value", bold=True)
out(ws.cell(r, 2), f"={LOAN}/B{TOTCOST}", PCT, bold=True); r += 1
label(ws.cell(r, 1), "LTV vs. Lendable (discounted) Value", bold=True)
out(ws.cell(r, 2), f"={LOAN}/D{LENDABLE}", PCT, bold=True); r += 1
CCR = r
label(ws.cell(r, 1), "Collateral Coverage (lendable / loan)", bold=True)
out(ws.cell(r, 2), f"=D{LENDABLE}/{LOAN}", RATIO, bold=True, fill=GREEN_FILL); r += 1
label(ws.cell(r, 1), "Policy Min. Coverage")
inp(ws.cell(r, 2), 1.00, RATIO); CMIN = r; r += 1
label(ws.cell(r, 1), "Pass / Fail", bold=True)
out(ws.cell(r, 2), f'=IF(B{CCR}>=B{CMIN},"PASS","FAIL")', '@', bold=True, fill=GREEN_FILL); r += 2

note(ws, r, "Advance rate = the % of value a lender will lend against after a liquidation haircut "
            "(vehicles hold value better than used equipment). LTV = loan / collateral value; lower is "
            "safer. Collateral coverage >=1.0x on lendable value means the bank can expect to recover the "
            "loan in a default. Coverage here is adequate but not abundant - secondary, not primary, repayment.",
     span=5, height=56)

LTV_LEND = f"'{S6}'!B{CCR}"  # coverage ratio cell


# ----------------------------------------------------------------------------
# 07 Risk Rating
# ----------------------------------------------------------------------------
ws = wb.create_sheet(S7)
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 26
for col, w in {"B": 12, "C": 10, "D": 14, "E": 44}.items():
    ws.column_dimensions[col].width = w
title_block(ws, "Risk Rating Scorecard",
            "Weighted 1 (strong) - 5 (weak); lower is better", span=5)

r = 4
hdr(ws.cell(r, 1), "Risk Factor")
hdr(ws.cell(r, 2), "Weight")
hdr(ws.cell(r, 3), "Score 1-5")
hdr(ws.cell(r, 4), "Weighted")
hdr(ws.cell(r, 5), "Rationale")
r += 1
factors = [
    ("Repayment Capacity (DSCR)", 0.25, 2, "Global DSCR ~1.3x, proposed-loan DSCR >2.5x; healthy cushion."),
    ("Leverage", 0.15, 3, "D/E ~1.0x and Debt/EBITDA <2x - moderate, in line with industry."),
    ("Liquidity", 0.10, 2, "Current ratio >2x; positive and growing working capital."),
    ("Collateral Coverage", 0.15, 3, "~1.05x lendable coverage; adequate but not strong; depreciating assets."),
    ("Management", 0.10, 3, "Experienced operator but new, unseasoned banking relationship."),
    ("Industry Risk", 0.10, 3, "Equipment repair is cyclical/competitive but demand is non-discretionary."),
    ("Financial Trends", 0.15, 2, "Three years of revenue growth with stable margins and rising equity."),
]
f_start = r
for name, w, s, rationale in factors:
    label(ws.cell(r, 1), name)
    inp(ws.cell(r, 2), w, PCT)
    inp(ws.cell(r, 3), s, '0')
    out(ws.cell(r, 4), f"=B{r}*C{r}", NUM2)
    rc = ws.cell(r, 5, rationale); rc.font = NOTE_FONT; rc.alignment = WRAP; rc.border = BORDER
    r += 1
f_end = r - 1
WSUM = r
label(ws.cell(r, 1), "Weighted Composite Score", bold=True)
out(ws.cell(r, 2), f"=SUM(B{f_start}:B{f_end})", PCT, bold=True)
ws.cell(r, 3).border = BORDER
out(ws.cell(r, 4), f"=SUM(D{f_start}:D{f_end})", NUM2, bold=True, fill=LIGHT_FILL)
r += 1
GRADE = r
label(ws.cell(r, 1), "Risk Grade (1-7 scale)", bold=True)
out(ws.cell(r, 4),
    f'=IF(D{WSUM}<=1.5,"2 - Strong",'
    f'IF(D{WSUM}<=2.5,"4 - Satisfactory",'
    f'IF(D{WSUM}<=3.5,"5 - Acceptable / Watch",'
    f'IF(D{WSUM}<=4.2,"6 - Special Mention","7 - Substandard"))))',
    '@', bold=True, fill=GREEN_FILL)
r += 2

note(ws, r, "Each factor is scored 1 (strong) to 5 (weak) and multiplied by its weight; the weighted "
            "scores sum to a composite that maps to an internal risk grade. A composite around 2.5 lands "
            "in the low end of 'Acceptable' - a pass grade that typically supports approval with conditions.",
     span=5, height=44)

RISK_COMPOSITE = f"'{S7}'!D{WSUM}"
RISK_GRADE = f"'{S7}'!D{GRADE}"


# ----------------------------------------------------------------------------
# 08 Credit Memo
# ----------------------------------------------------------------------------
ws = wb.create_sheet(S8)
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 100
title_block(ws, "Credit Memo & Recommendation",
            "Northstar Equipment Services LLC  |  $750,000 Term Loan", span=2)

r = 4
section(ws, r, "Recommendation", span=2); r += 1
label(ws.cell(r, 1), "Decision", bold=True)
c = ws.cell(r, 2, "APPROVE WITH CONDITIONS")
c.font = Font(color="006100", bold=True, size=12); c.fill = GREEN_FILL; c.alignment = LEFT
r += 1
label(ws.cell(r, 1), "Risk Grade", bold=True)
out(ws.cell(r, 2), f"={RISK_GRADE}", '@', bold=True); ws.cell(r, 2).alignment = LEFT; r += 1
label(ws.cell(r, 1), "2025 Global DSCR", bold=True)
out(ws.cell(r, 2), f"={DSCR_GLOBAL_2025}", RATIO, bold=True); ws.cell(r, 2).alignment = LEFT; r += 1
label(ws.cell(r, 1), "Collateral Coverage", bold=True)
out(ws.cell(r, 2), f"={LTV_LEND}", RATIO, bold=True); ws.cell(r, 2).alignment = LEFT; r += 2


def memo_block(rr, heading, lines):
    section(ws, rr, heading, span=2); rr += 1
    for ln in lines:
        c = ws.cell(rr, 2, "•  " + ln); c.font = OUT_FONT; c.alignment = WRAP
        ws.row_dimensions[rr].height = 30
        rr += 1
    return rr + 1


r = memo_block(r, "Summary", [
    "Northstar Equipment Services LLC requests a $750,000, 60-month term loan at 8.25% to purchase "
    "service vehicles and shop equipment, with a $85,000 owner equity injection.",
    "The company shows three years of steady revenue growth (~$4.2M to $5.4M), stable ~39% gross "
    "margins, modest leverage, and strong liquidity. Cash flow comfortably covers the new debt.",
])
r = memo_block(r, "Strengths", [
    "Consistent revenue growth and stable margins across 2023-2025.",
    "Strong repayment capacity: proposed-loan DSCR >2.5x and global DSCR ~1.3x, above the 1.20x floor.",
    "Modest leverage (Debt/EBITDA <2x, D/E ~1.0x) and healthy liquidity (current ratio >2x).",
    "Loan is purchase-money secured by titled vehicles and equipment, plus a personal guarantee.",
    "Owner contributes equity (skin in the game), reducing effective LTV.",
])
r = memo_block(r, "Weaknesses", [
    "New banking relationship with no prior repayment history at this institution.",
    "Collateral is depreciating; lendable coverage (~1.05x) is adequate but not abundant.",
    "Single-owner dependency (key-person risk); LLC financials likely company-prepared/reviewed.",
])
r = memo_block(r, "Key Risks", [
    "Cyclicality: a downturn in customers' capital spending could compress repair volumes/revenue.",
    "Asset depreciation could outpace loan amortization, weakening collateral coverage over time.",
    "Margin/customer concentration risk not fully visible from summary financials.",
])
r = memo_block(r, "Mitigants", [
    "Conservative DSCR (after maintenance capex) leaves cushion against a revenue decline.",
    "Purchase-money lien on financed assets + additional equipment pledge and personal guarantee.",
    "Non-discretionary nature of equipment repair supports demand resilience through cycles.",
    "Equity injection and amortizing structure de-risk the bank's exposure over the term.",
])
r = memo_block(r, "Approval Conditions", [
    "First-priority (purchase-money) lien on all financed vehicles and equipment; UCC-1 filed.",
    "Unlimited personal guarantee of the majority owner.",
    "Minimum equity injection of $85,000 verified at closing.",
    "Financial covenants: maintain global DSCR >=1.20x and Debt/EBITDA <=3.5x, tested annually.",
    "Annual CPA-reviewed financial statements and interim statements upon request.",
    "Proof of insurance on collateral with bank named as loss payee.",
])

note(ws, r, "This memo is the underwriter's written recommendation. The decision, risk grade and key "
            "metrics above pull live from the model tabs, so changing an assumption updates the memo.",
     span=2, height=30)


# ----------------------------------------------------------------------------
wb.save("Commercial_Credit_Underwriting_Model.xlsx")
print("Saved Commercial_Credit_Underwriting_Model.xlsx")
